import requests
import pandas as pd
import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime  # Import for timestamp conversion
import threading

# Deezer API URL for fetching playlist tracks
DEEZER_API_URL = "https://api.deezer.com/"

# Fetch the playlist data using the ARL token and handles pagination
def get_playlist_data(arl_token, playlist_id, callback):
    headers = {
        "Authorization": f"Bearer {arl_token}"
    }
    
    # URL for the first page of playlist tracks
    playlist_url = f"{DEEZER_API_URL}playlist/{playlist_id}/tracks"
    
    # List to store all tracks
    all_tracks = []
    
    while playlist_url:
        # Fetch the playlist details
        response = requests.get(playlist_url, headers=headers)
        
        # Check if the response is successful
        if response.status_code != 200:
            messagebox.showerror("Error", f"Error fetching playlist data: {response.status_code} - {response.text}")
            callback(None)  # Call the callback function with None to indicate an error
            return

        # Parse the response as JSON
        playlist_data = response.json()
        
        # Check if the 'data' key exists in the response
        if 'data' not in playlist_data:
            messagebox.showerror("Error", f"Unexpected response format: {playlist_data}")
            callback(None)  # Call the callback function with None to indicate an error
            return
        
        # Add the current page's tracks to the list
        all_tracks.extend(playlist_data["data"])
        
        # Check if there is a next page by safely getting the 'next' key
        playlist_url = playlist_data.get("next")  # Using .get() will safely return None if 'next' is not found
    
    # Once all tracks are collected, call the callback to process the data
    callback(all_tracks)

# Convert playlist data to a pandas DataFrame
def create_dataframe_from_playlist(playlist_data):
    tracks = []
    
    for idx, track in enumerate(playlist_data, start=1):
        # Convert Unix timestamp to readable date format, if available
        date_added = datetime.fromtimestamp(track["time_add"]).strftime('%Y-%m-%d %H:%M:%S') if "time_add" in track else "Unknown"
        
        track_info = {
            "Track Number": idx,  # Add track number based on position in playlist
            "Track Name": track["title"],
            "Artist": track["artist"]["name"],
            "Album": track["album"]["title"],
            "Duration": track["duration"],
            "Link": track["link"],
            "Date Added": date_added  # New Date Added column
        }
        tracks.append(track_info)
    
    df = pd.DataFrame(tracks)
    return df

# Save DataFrame to Excel and apply formatting
def save_to_excel(df, file_path):
    # Save DataFrame to Excel using pandas
    df.to_excel(file_path, index=False, engine='openpyxl')
    
    # Load the workbook to apply formatting
    wb = load_workbook(file_path)
    
    # Apply formatting to all worksheets
    for ws in wb.worksheets:
        # Step 1: Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Approximate auto-fit
            ws.column_dimensions[column].width = adjusted_width
        
        # Step 2: Cap column width at 50 and add space to adjacent empty cells
        for col in ws.columns:
            col_letter = col[0].column_letter
            col_width = ws.column_dimensions[col_letter].width or 0
            if col_width > 30:
                ws.column_dimensions[col_letter].width = 50
                # Add space to adjacent column if it exists and is empty
                col_idx = col[0].column
                next_col_letter = get_column_letter(col_idx + 1)
                if next_col_letter in ws.column_dimensions:
                    for row in range(1, ws.max_row + 1):
                        cell = ws[f"{next_col_letter}{row}"]
                        if not cell.value:
                            cell.value = " "
        
        # Step 3: Freeze the top row
        ws.freeze_panes = ws['A2']  # Freeze below the first row
    
    # Save the formatted workbook
    wb.save(file_path)
    messagebox.showinfo("Success", f"Playlist data saved to {file_path} with formatting")

# Function to run the script in a background thread
def run_script():
    arl_token = entry_arl_token.get()
    playlist_id = entry_playlist_id.get()
    file_path = entry_file_path.get()
    
    if not arl_token or not playlist_id or not file_path:
        messagebox.showerror("Error", "Please fill all fields!")
        return

    # Disable the "Export Playlist" button to prevent multiple clicks
    btn_run.config(state=tk.DISABLED)
    
    # Start the background thread to fetch playlist data
    threading.Thread(target=get_playlist_data, args=(arl_token, playlist_id, process_playlist_data)).start()

# Callback function to process the playlist data after the background task finishes
def process_playlist_data(playlist_data):
    # Enable the "Export Playlist" button again
    btn_run.config(state=tk.NORMAL)
    
    if playlist_data is None:
        return  # If there was an error, just return

    # Process and save the data
    df = create_dataframe_from_playlist(playlist_data)
    save_to_excel(df, entry_file_path.get())

# Function to browse for file path
def browse_file():
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        entry_file_path.delete(0, tk.END)
        entry_file_path.insert(0, file_path)

# Create the GUI window
root = tk.Tk()
root.title("Deezer Playlist Exporter")

# Set the window size
window_width = 400
window_height = 300

# Get the screen width and height
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Calculate the position to center the window
x = (screen_width // 2) - (window_width // 2)
y = (screen_height // 2) - (window_height // 2)

# Set the window geometry with calculated position
root.geometry(f"{window_width}x{window_height}+{x}+{y}")

# Labels and inputs for ARL token, Playlist ID, and File path
tk.Label(root, text="ARL Token:").pack(pady=5)
entry_arl_token = tk.Entry(root, width=40)
entry_arl_token.pack(pady=5)

tk.Label(root, text="Playlist ID:").pack(pady=5)
entry_playlist_id = tk.Entry(root, width=40)
entry_playlist_id.pack(pady=5)

tk.Label(root, text="Save as (Excel file):").pack(pady=5)
entry_file_path = tk.Entry(root, width=40)
entry_file_path.pack(pady=5)

# Browse button for file path
btn_browse = tk.Button(root, text="Browse", command=browse_file)
btn_browse.pack(pady=5)

# Run button
btn_run = tk.Button(root, text="Export Playlist", command=run_script)
btn_run.pack(pady=20)

# Run the GUI main loop
root.mainloop()
